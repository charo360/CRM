# zilo_excel_styles.py

import re
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.utils import get_column_letter

# ── COLORS ──
DARK="0A0A0A"; GREEN="22C55E"; WHITE="FFFFFF"
LGRAY="F9FAFB"; MGRAY="E5E7EB"; AMBER_L="FFFFF3CD"
GREEN_L="D1FAE5"; GREEN_D="166534"; AMBER="F59E0B"
RED_L="FEE2E2"; RED_D="991B1B"; DGRAY="374151"
BLUE="0000FF"; BLACK="000000"; GREEN_TX="008000"
MUTED="9CA3AF"; CARD="111111"

def fill(hex_color): 
    # Clean hash if present
    hex_clean = hex_color.lstrip("#").upper()
    return PatternFill("solid", fgColor=hex_clean)

def border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s,right=s,top=s,bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(
        horizontal=h, vertical=v, wrap_text=wrap
    )

def _is_light_color(hex_color: str) -> bool:
    """Check if hex color is light to determine text contrast."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 3:
        hex_color = "".join(c*2 for c in hex_color)
    if len(hex_color) != 6:
        return True
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except ValueError:
        return True
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return luminance > 180

def _get_tint_color(hex_color: str, factor: float) -> str:
    """Mix the hex color with white. factor = 0.0 is pure white, 1.0 is original color."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 3:
        hex_color = "".join(c*2 for c in hex_color)
    if len(hex_color) != 6:
        return "FFFFFF"
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except ValueError:
        return "FFFFFF"
    r = int(r * factor + 255 * (1 - factor))
    g = int(g * factor + 255 * (1 - factor))
    b = int(b * factor + 255 * (1 - factor))
    return f"{r:02X}{g:02X}{b:02X}"

def title_band(ws, row, title, subtitle="", cols=14):
    ws.row_dimensions[row].height = 44
    # Detect primary brand color from workbook if set
    brand_color = getattr(ws.parent, "primary_color", DARK)
    text_color = "000000" if _is_light_color(brand_color) else "FFFFFF"
    
    for col in range(1, cols+1):
        ws.cell(row,col).fill = fill(brand_color)
    c = ws.cell(row,1)
    c.value = title
    c.font = Font(name="Arial",bold=True,
                  size=16,color=text_color)
    c.alignment = align("left","center")
    if subtitle:
        cs = ws.cell(row, cols-1)
        cs.value = subtitle
        cs.font = Font(name="Arial",size=9,
                       color=MUTED)
        cs.alignment = align("right","center")

def section_hdr(ws, row, number, title, cols=14):
    ws.row_dimensions[row].height = 22
    brand_color = getattr(ws.parent, "primary_color", GREEN)
    text_color = "000000" if _is_light_color(brand_color) else "FFFFFF"
    
    for col in range(1, cols+1):
        ws.cell(row,col).fill = fill(brand_color)
    c = ws.cell(row,1)
    c.value = f"  {number:02d}  {title}"
    c.font = Font(name="Arial",bold=True,
                  size=9,color=text_color)
    c.alignment = align("left","center")

def col_hdr(ws, row, headers, bg=CARD):
    ws.row_dimensions[row].height = 20
    brand_color = getattr(ws.parent, "primary_color", bg)
    text_color = "000000" if _is_light_color(brand_color) else "FFFFFF"
    
    for i,h in enumerate(headers,1):
        c = ws.cell(row,i)
        c.value = h
        c.font = Font(name="Arial",bold=True,
                      size=9,color=text_color)
        c.fill = fill(brand_color)
        c.alignment = align("center","center")
        c.border = border()

def label(ws, row, col, value,
          bg=None, bold=False, indent=False):
    r = row % 2
    # Mix dynamic brand color at 3% for alternating rows if primary brand color is set
    brand_color = getattr(ws.parent, "primary_color", None)
    if brand_color and not bg:
        alt_bg = _get_tint_color(brand_color, 0.03)
        bg_color = alt_bg if r==0 else WHITE
    else:
        bg_color = bg or (LGRAY if r==0 else WHITE)
        
    ws.row_dimensions[row].height = 17
    c = ws.cell(row,col)
    c.value = ("    " if indent else "") + str(value)
    c.font = Font(name="Arial",size=9,
                  bold=bold,color=DGRAY)
    c.fill = fill(bg_color)
    c.alignment = align("left","center")
    c.border = border()

def value(ws, row, col, val, fmt="",
          input_type="input", bg=None,
          bold=False):
    r = row % 2
    # Mix dynamic brand color at 3% for alternating rows if primary brand color is set
    brand_color = getattr(ws.parent, "primary_color", None)
    if brand_color and not bg:
        alt_bg = _get_tint_color(brand_color, 0.03)
        bg_color = alt_bg if r==0 else WHITE
    else:
        bg_color = bg or (LGRAY if r==0 else WHITE)
        
    ws.row_dimensions[row].height = 17
    color_map = {
        "input":    BLUE,
        "formula":  BLACK,
        "link":     GREEN_TX,
        "text":     DGRAY,
        "total":    DGRAY,
    }
    c = ws.cell(row,col)
    c.value = val
    c.font = Font(name="Arial",size=9,bold=bold,
                  color=color_map.get(input_type,BLACK))
    if fmt: c.number_format = fmt
    c.fill = fill(bg_color)
    c.alignment = align("center","center")
    c.border = border()

def total_row(ws, row, label_text, 
              formulas, fmt, cols=14):
    ws.row_dimensions[row].height = 20
    
    brand_color = getattr(ws.parent, "primary_color", None)
    # Mix dynamic brand color at 8% for total row if primary brand color is set
    bg_color = _get_tint_color(brand_color, 0.08) if brand_color else AMBER_L
    
    c = ws.cell(row,1)
    c.value = label_text
    c.font = Font(name="Arial",bold=True,
                  size=9,color=DGRAY)
    c.fill = fill(bg_color)
    c.alignment = align("left","center")
    c.border = border()
    for i,(col,formula) in enumerate(
            zip(range(2,cols+1),formulas)):
        cv = ws.cell(row,col)
        cv.value = formula
        cv.font = Font(name="Arial",bold=True,
                       size=9,color=DGRAY)
        cv.number_format = fmt
        cv.fill = fill(bg_color)
        cv.alignment = align("center","center")
        cv.border = border()

def gap(ws, row):
    ws.row_dimensions[row].height = 6

def set_col(ws, col, width):
    ws.column_dimensions[
        get_column_letter(col)
    ].width = width

def setup_sheet(ws, freeze="B4"):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_view.zoomScale = 100

def tab_color(ws, hex):
    ws.sheet_properties.tabColor = hex.lstrip("#").upper()
